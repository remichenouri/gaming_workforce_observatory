"""
Gaming Workforce Observatory - Export Manager Enterprise
Gestionnaire d'export avancé PDF/Excel avec templates gaming
"""
import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Tuple
import streamlit as st
from datetime import datetime
import io
import base64
from pathlib import Path
import logging
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.colors import Color, HexColor
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.drawing.image import Image as XLImage

logger = logging.getLogger(__name__)

class GamingExportManager:
    """Gestionnaire d'export enterprise pour rapports gaming workforce"""
    
    def __init__(self):
        # Configuration couleurs gaming/Ubisoft
        self.colors = {
            'ubisoft_blue': HexColor('#0082c4'),
            'ubisoft_cyan': HexColor('#00d4ff'),
            'gaming_orange': HexColor('#ff6b35'),
            'success_green': HexColor('#00ff88'),
            'warning_yellow': HexColor('#ffaa00'),
            'danger_red': HexColor('#ff3366'),
            'dark_bg': HexColor('#1a1a2e'),
            'light_gray': HexColor('#ecf0f1')
        }
        
        # Styles PDF gaming
        self.pdf_styles = {
            'title': ParagraphStyle(
                'GameTitle',
                fontName='Helvetica-Bold',
                fontSize=24,
                textColor=self.colors['ubisoft_blue'],
                spaceAfter=20,
                alignment=1  # CENTER
            ),
            'heading': ParagraphStyle(
                'GameHeading',
                fontName='Helvetica-Bold',
                fontSize=16,
                textColor=self.colors['ubisoft_cyan'],
                spaceAfter=12,
                spaceBefore=12
            ),
            'body': ParagraphStyle(
                'GameBody',
                fontName='Helvetica',
                fontSize=10,
                textColor=HexColor('#2c3e50'),
                spaceAfter=6,
                leading=12
            )
        }
        
        # Templates Excel gaming
        self.excel_styles = {
            'header': {
                'font': Font(name='Calibri', size=14, bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='0082C4', end_color='0082C4', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'subheader': {
                'font': Font(name='Calibri', size=12, bold=True, color='0082C4'),
                'fill': PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center')
            },
            'data': {
                'font': Font(name='Calibri', size=10),
                'alignment': Alignment(horizontal='left', vertical='center')
            },
            'metric': {
                'font': Font(name='Calibri', size=12, bold=True, color='FF6B35'),
                'alignment': Alignment(horizontal='center', vertical='center')
            }
        }
    
    def export_executive_summary_pdf(self, summary_data: Dict[str, Any], 
                                   filename: str = None) -> bytes:
        """Exporte un résumé exécutif en PDF gaming"""
        
        if filename is None:
            filename = f"gaming_workforce_executive_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        # Buffer pour PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1*inch)
        
        # Contenus du PDF
        story = []
        
        # Titre principal
        title = Paragraph("🎮 Gaming Workforce Observatory", self.pdf_styles['title'])
        story.append(title)
        story.append(Spacer(1, 0.5*inch))
        
        # Sous-titre avec date
        subtitle = Paragraph(f"Executive Summary - {datetime.now().strftime('%B %d, %Y')}", 
                            self.pdf_styles['heading'])
        story.append(subtitle)
        story.append(Spacer(1, 0.3*inch))
        
        # Métriques clés
        if 'key_metrics' in summary_data:
            story.append(Paragraph("📊 Key Performance Indicators", self.pdf_styles['heading']))
            
            metrics_data = [
                ['Metric', 'Current Value', 'Target', 'Status'],
            ]
            
            for metric_name, metric_info in summary_data['key_metrics'].items():
                current = metric_info.get('current', 'N/A')
                target = metric_info.get('target', 'N/A')
                status = metric_info.get('status', 'Unknown')
                
                metrics_data.append([
                    metric_name.replace('_', ' ').title(),
                    str(current),
                    str(target),
                    status
                ])
            
            metrics_table = Table(metrics_data, colWidths=[3*inch, 1.5*inch, 1.5*inch, 1*inch])
            metrics_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), self.colors['ubisoft_blue']),
                ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#FFFFFF')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), self.colors['light_gray']),
                ('GRID', (0, 0), (-1, -1), 1, self.colors['ubisoft_blue'])
            ]))
            
            story.append(metrics_table)
            story.append(Spacer(1, 0.3*inch))
        
        # Analyse des départements
        if 'department_analysis' in summary_data:
            story.append(Paragraph("🏢 Department Analysis", self.pdf_styles['heading']))
            
            for dept, analysis in summary_data['department_analysis'].items():
                dept_title = Paragraph(f"• {dept}", self.pdf_styles['body'])
                story.append(dept_title)
                
                employee_count = analysis.get('employee_count', 0)
                avg_satisfaction = analysis.get('avg_satisfaction', 0)
                retention_rate = analysis.get('retention_rate', 0)
                
                dept_info = Paragraph(
                    f"Employees: {employee_count} | "
                    f"Satisfaction: {avg_satisfaction:.1f}/10 | "
                    f"Retention: {retention_rate:.1f}%",
                    self.pdf_styles['body']
                )
                story.append(dept_info)
                story.append(Spacer(1, 0.1*inch))
        
        # Recommandations
        if 'recommendations' in summary_data:
            story.append(Paragraph("💡 Strategic Recommendations", self.pdf_styles['heading']))
            
            for i, recommendation in enumerate(summary_data['recommendations'][:5], 1):
                rec_text = f"{i}. {recommendation.get('title', 'Recommendation')}: " \
                          f"{recommendation.get('description', 'No description')}"
                rec_para = Paragraph(rec_text, self.pdf_styles['body'])
                story.append(rec_para)
                story.append(Spacer(1, 0.1*inch))
        
        # Footer
        story.append(Spacer(1, 0.5*inch))
        footer = Paragraph(
            "Generated by Gaming Workforce Observatory Enterprise | Confidential",
            ParagraphStyle('Footer', fontName='Helvetica-Oblique', fontSize=8, 
                          textColor=HexColor('#7f8c8d'), alignment=1)
        )
        story.append(footer)
        
        # Construction du PDF
        doc.build(story)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def export_detailed_workforce_excel(self, workforce_data: Dict[str, pd.DataFrame],
                                      filename: str = None) -> bytes:
        """Exporte un rapport workforce détaillé en Excel"""
        
        if filename is None:
            filename = f"gaming_workforce_detailed_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Buffer pour Excel
        buffer = io.BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            
            # Feuille 1: Dashboard Overview
            self._create_dashboard_sheet(writer, workforce_data)
            
            # Feuille 2: Employee Details
            if 'employees' in workforce_data:
                self._create_employee_details_sheet(writer, workforce_data['employees'])
            
            # Feuille 3: Department Analysis
            if 'departments' in workforce_data:
                self._create_department_analysis_sheet(writer, workforce_data['departments'])
            
            # Feuille 4: Skills Matrix
            if 'skills' in workforce_data:
                self._create_skills_matrix_sheet(writer, workforce_data['skills'])
            
            # Feuille 5: Performance Metrics
            if 'performance' in workforce_data:
                self._create_performance_metrics_sheet(writer, workforce_data['performance'])
        
        buffer.seek(0)
        return buffer.getvalue()
    
    def _create_dashboard_sheet(self, writer: pd.ExcelWriter, data: Dict[str, pd.DataFrame]):
        """Crée la feuille dashboard overview"""
        
        workbook = writer.book
        worksheet = workbook.create_sheet('Dashboard Overview', 0)
        
        # Titre principal
        worksheet['A1'] = '🎮 Gaming Workforce Observatory - Dashboard'
        worksheet['A1'].font = Font(name='Calibri', size=18, bold=True, color='0082C4')
        worksheet.merge_cells('A1:F1')
        worksheet['A1'].alignment = Alignment(horizontal='center')
        
        # Date de génération
        worksheet['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        worksheet['A2'].font = Font(name='Calibri', size=10, italic=True)
        worksheet.merge_cells('A2:F2')
        
        # Métriques clés
        row = 4
        worksheet[f'A{row}'] = 'KEY METRICS'
        worksheet[f'A{row}'].font = self.excel_styles['header']['font']
        worksheet[f'A{row}'].fill = self.excel_styles['header']['fill']
        worksheet.merge_cells(f'A{row}:D{row}')
        
        row += 1
        metrics = [
            ('Total Employees', len(data.get('employees', pd.DataFrame()))),
            ('Departments', len(data.get('departments', pd.DataFrame()))),
            ('Active Projects', 23),  # Exemple
            ('Avg Satisfaction', 7.2)  # Exemple
        ]
        
        for i, (metric_name, value) in enumerate(metrics):
            col = chr(65 + i)  # A, B, C, D
            worksheet[f'{col}{row}'] = metric_name
            worksheet[f'{col}{row}'].font = self.excel_styles['subheader']['font']
            worksheet[f'{col}{row}'].fill = self.excel_styles['subheader']['fill']
            
            worksheet[f'{col}{row+1}'] = value
            worksheet[f'{col}{row+1}'].font = self.excel_styles['metric']['font']
            worksheet[f'{col}{row+1}'].alignment = self.excel_styles['metric']['alignment']
        
        # Départements overview
        if 'employees' in data and not data['employees'].empty:
            row += 4
            worksheet[f'A{row}'] = 'DEPARTMENT BREAKDOWN'
            worksheet[f'A{row}'].font = self.excel_styles['header']['font']
            worksheet[f'A{row}'].fill = self.excel_styles['header']['fill']
            worksheet.merge_cells(f'A{row}:C{row}')
            
            row += 1
            dept_counts = data['employees'].groupby('department').size().reset_index(name='count')
            
            headers = ['Department', 'Employee Count', 'Percentage']
            for i, header in enumerate(headers):
                col = chr(65 + i)
                worksheet[f'{col}{row}'] = header
                worksheet[f'{col}{row}'].font = self.excel_styles['subheader']['font']
                worksheet[f'{col}{row}'].fill = self.excel_styles['subheader']['fill']
            
            row += 1
            total_employees = len(data['employees'])
            
            for _, dept_row in dept_counts.iterrows():
                worksheet[f'A{row}'] = dept_row['name'] if 'name' in dept_row else 'Unknown'
                worksheet[f'B{row}'] = dept_row['count']
                worksheet[f'C{row}'] = f"{(dept_row['count']/total_employees)*100:.1f}%"
                
                for col in ['A', 'B', 'C']:
                    worksheet[f'{col}{row}'].font = self.excel_styles['data']['font']
                    worksheet[f'{col}{row}'].alignment = self.excel_styles['data']['alignment']
                
                row += 1
        
        # Ajustement largeur colonnes
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            worksheet.column_dimensions[col].width = 20
    
    def _create_employee_details_sheet(self, writer: pd.ExcelWriter, employees_df: pd.DataFrame):
        """Crée la feuille détails employés"""
        
        # Export du DataFrame avec styles
        employees_df.to_excel(writer, sheet_name='Employee Details', index=False)
        
        workbook = writer.book
        worksheet = workbook['Employee Details']
        
        # Style de l'en-tête
        for cell in worksheet[1]:
            cell.font = self.excel_styles['header']['font']
            cell.fill = self.excel_styles['header']['fill']
            cell.alignment = self.excel_styles['header']['alignment']
        
        # Style des données
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = self.excel_styles['data']['font']
                cell.alignment = self.excel_styles['data']['alignment']
        
        # Ajustement largeur colonnes
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _create_department_analysis_sheet(self, writer: pd.ExcelWriter, departments_df: pd.DataFrame):
        """Crée la feuille analyse départements avec graphiques"""
        
        departments_df.to_excel(writer, sheet_name='Department Analysis', index=False)
        
        workbook = writer.book
        worksheet = workbook['Department Analysis']
        
        # Style de base
        self._apply_basic_excel_styling(worksheet)
        
        # Ajout d'un graphique si les données le permettent
        if 'employee_count' in departments_df.columns and len(departments_df) > 1:
            
            # Données pour le graphique
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = "Employees by Department"
            chart.y_axis.title = 'Number of Employees'
            chart.x_axis.title = 'Department'
            
            # Références des données
            data = Reference(worksheet, min_col=2, min_row=1, max_row=len(departments_df)+1, max_col=2)
            cats = Reference(worksheet, min_col=1, min_row=2, max_row=len(departments_df)+1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Positionnement du graphique
            worksheet.add_chart(chart, "E2")
    
    def _create_skills_matrix_sheet(self, writer: pd.ExcelWriter, skills_df: pd.DataFrame):
        """Crée la feuille matrice des compétences"""
        
        skills_df.to_excel(writer, sheet_name='Skills Matrix', index=False)
        
        workbook = writer.book
        worksheet = workbook['Skills Matrix']
        
        self._apply_basic_excel_styling(worksheet)
        
        # Formatage conditionnel pour les niveaux de compétences
        from openpyxl.formatting.rule import ColorScaleRule
        
        # Trouve les colonnes de compétences (supposées numériques)
        skill_cols = []
        for i, cell in enumerate(worksheet[1], 1):
            if 'skill' in str(cell.value).lower() or 'level' in str(cell.value).lower():
                skill_cols.append(i)
        
        if skill_cols:
            for col_idx in skill_cols:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                range_string = f"{col_letter}2:{col_letter}{len(skills_df)+1}"
                
                # Échelle de couleur rouge-jaune-vert
                rule = ColorScaleRule(
                    start_type='min', start_color='FF6B35',  # Orange/Rouge
                    mid_type='percentile', mid_value=50, mid_color='FFAA00',  # Jaune
                    end_type='max', end_color='00FF88'  # Vert
                )
                
                worksheet.conditional_formatting.add(range_string, rule)
    
    def _create_performance_metrics_sheet(self, writer: pd.ExcelWriter, performance_df: pd.DataFrame):
        """Crée la feuille métriques de performance"""
        
        performance_df.to_excel(writer, sheet_name='Performance Metrics', index=False)
        
        workbook = writer.book
        worksheet = workbook['Performance Metrics']
        
        self._apply_basic_excel_styling(worksheet)
        
        # Graphique de performance si possible
        if 'performance_score' in performance_df.columns and len(performance_df) > 5:
            
            # Histogramme des scores de performance
            chart = BarChart()
            chart.title = "Performance Score Distribution"
            chart.y_axis.title = 'Count'
            chart.x_axis.title = 'Performance Score'
            
            # Calcul de la distribution
            bins = np.histogram(performance_df['performance_score'], bins=5)
            
            # Ajout des données de distribution dans des cellules temporaires
            start_row = len(performance_df) + 5
            for i, (count, bin_edge) in enumerate(zip(bins[0], bins[1][:-1])):
                worksheet[f'A{start_row + i}'] = f"{bin_edge:.1f}-{bins[1][i+1]:.1f}"
                worksheet[f'B{start_row + i}'] = count
            
            # Références pour le graphique
            data = Reference(worksheet, min_col=2, min_row=start_row, max_row=start_row+len(bins[0])-1)
            cats = Reference(worksheet, min_col=1, min_row=start_row, max_row=start_row+len(bins[0])-1)
            
            chart.add_data(data)
            chart.set_categories(cats)
            
            worksheet.add_chart(chart, f"D{start_row}")
    
    def _apply_basic_excel_styling(self, worksheet):
        """Applique le style de base Excel gaming"""
        
        # Style de l'en-tête
        for cell in worksheet[1]:
            cell.font = self.excel_styles['header']['font']
            cell.fill = self.excel_styles['header']['fill']
            cell.alignment = self.excel_styles['header']['alignment']
        
        # Style des données
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = self.excel_styles['data']['font']
                cell.alignment = self.excel_styles['data']['alignment']
        
        # Ajustement automatique des colonnes
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def create_download_link(self, file_bytes: bytes, filename: str, 
                           file_type: str = "application/octet-stream") -> str:
        """Crée un lien de téléchargement Streamlit"""
        
        b64 = base64.b64encode(file_bytes).decode()
        
        return f'<a href="data:{file_type};base64,{b64}" download="{filename}" ' \
               f'style="background: linear-gradient(45deg, #0082c4, #00d4ff); ' \
               f'color: white; padding: 10px 20px; text-decoration: none; ' \
               f'border-radius: 5px; font-weight: bold;">📥 Download {filename}</a>'
    
    def render_export_interface(self, data_dict: Dict[str, Any]):
        """Interface Streamlit pour les exports"""
        
        st.markdown("## 📤 Export & Reporting Center")
        st.markdown("*Generate professional reports and exports for gaming workforce data*")
        
        # Sélection du type d'export
        export_type = st.selectbox(
            "Choose Export Type:",
            ["Executive Summary (PDF)", "Detailed Workforce Report (Excel)", 
             "Department Analysis (Excel)", "Custom Data Export"]
        )
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Options d'export
            if export_type == "Executive Summary (PDF)":
                st.markdown("### 📋 Executive Summary Options")
                
                include_metrics = st.checkbox("Include Key Metrics", value=True)
                include_departments = st.checkbox("Include Department Analysis", value=True)
                include_recommendations = st.checkbox("Include Recommendations", value=True)
                
                if st.button("Generate PDF Report"):
                    with st.spinner("Generating executive summary..."):
                        
                        # Préparation des données résumé
                        summary_data = {
                            'key_metrics': {
                                'total_employees': {'current': 850, 'target': 900, 'status': 'On Track'},
                                'avg_satisfaction': {'current': 7.2, 'target': 8.0, 'status': 'Needs Improvement'},
                                'retention_rate': {'current': 85.5, 'target': 90.0, 'status': 'Good'},
                                'project_completion': {'current': 92.0, 'target': 95.0, 'status': 'Good'}
                            } if include_metrics else {},
                            'department_analysis': {
                                'Programming': {'employee_count': 320, 'avg_satisfaction': 7.8, 'retention_rate': 88.0},
                                'Art & Animation': {'employee_count': 180, 'avg_satisfaction': 7.1, 'retention_rate': 82.0},
                                'Game Design': {'employee_count': 120, 'avg_satisfaction': 7.5, 'retention_rate': 90.0},
                                'Quality Assurance': {'employee_count': 150, 'avg_satisfaction': 6.8, 'retention_rate': 80.0},
                                'Production': {'employee_count': 80, 'avg_satisfaction': 7.3, 'retention_rate': 87.0}
                            } if include_departments else {},
                            'recommendations': [
                                {'title': 'Improve QA Satisfaction', 'description': 'Focus on QA team engagement and tools'},
                                {'title': 'Enhance Retention Programs', 'description': 'Implement targeted retention strategies'},
                                {'title': 'Expand Training Programs', 'description': 'Increase skill development opportunities'},
                                {'title': 'Review Compensation', 'description': 'Conduct market salary analysis'},
                                {'title': 'Strengthen Team Collaboration', 'description': 'Improve cross-department communication'}
                            ] if include_recommendations else []
                        }
                        
                        # Génération du PDF
                        pdf_bytes = self.export_executive_summary_pdf(summary_data)
                        
                        # Lien de téléchargement
                        filename = f"gaming_executive_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                        download_link = self.create_download_link(pdf_bytes, filename, "application/pdf")
                        
                        st.success("✅ Executive summary generated successfully!")
                        st.markdown(download_link, unsafe_allow_html=True)
            
            elif export_type == "Detailed Workforce Report (Excel)":
                st.markdown("### 📊 Workforce Report Options")
                
                include_employees = st.checkbox("Employee Details", value=True)
                include_departments = st.checkbox("Department Analysis", value=True)
                include_skills = st.checkbox("Skills Matrix", value=True)
                include_performance = st.checkbox("Performance Metrics", value=True)
                
                if st.button("Generate Excel Report"):
                    with st.spinner("Generating workforce report..."):
                        
                        # Données simulées pour la démo
                        workforce_data = {}
                        
                        if include_employees:
                            workforce_data['employees'] = pd.DataFrame({
                                'employee_id': range(1, 101),
                                'name': [f'Employee {i}' for i in range(1, 101)],
                                'department': np.random.choice(['Programming', 'Art & Animation', 'Game Design', 'QA', 'Production'], 100),
                                'experience_years': np.random.randint(1, 15, 100),
                                'salary_usd': np.random.randint(50000, 150000, 100),
                                'performance_score': np.random.uniform(2.0, 5.0, 100),
                                'satisfaction_score': np.random.uniform(4.0, 10.0, 100)
                            })
                        
                        if include_departments:
                            workforce_data['departments'] = pd.DataFrame({
                                'department': ['Programming', 'Art & Animation', 'Game Design', 'QA', 'Production'],
                                'employee_count': [320, 180, 120, 150, 80],
                                'avg_salary': [95000, 75000, 85000, 65000, 90000],
                                'avg_satisfaction': [7.8, 7.1, 7.5, 6.8, 7.3]
                            })
                        
                        # Génération de l'Excel
                        excel_bytes = self.export_detailed_workforce_excel(workforce_data)
                        
                        # Lien de téléchargement
                        filename = f"gaming_workforce_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                        download_link = self.create_download_link(
                            excel_bytes, filename, 
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        
                        st.success("✅ Workforce report generated successfully!")
                        st.markdown(download_link, unsafe_allow_html=True)
        
        with col2:
            # Prévisualisation et statistiques
            st.markdown("### 📈 Export Statistics")
            
            col1_stats, col2_stats = st.columns(2)
            
            with col1_stats:
                st.metric("📄 Reports Generated Today", "47", delta="12")
                st.metric("👥 Users Served", "23", delta="5")
            
            with col2_stats:
                st.metric("📊 Excel Exports", "28", delta="8")
                st.metric("📋 PDF Reports", "19", delta="4")
            
            # Formats supportés
            st.markdown("### 📁 Supported Formats")
            st.markdown("""
            **📋 PDF Reports:**
            - Executive Summaries
            - Department Analysis
            - Performance Reviews
            
            **📊 Excel Workbooks:**
            - Multi-sheet Reports
            - Interactive Charts
            - Conditional Formatting
            
            **📈 Data Formats:**
            - CSV for Raw Data
            - JSON for API Integration
            - Custom Templates Available
            """)
            
            # Historique récent
            st.markdown("### 🕒 Recent Exports")
            recent_exports = [
                "Executive_Summary_20250913_142301.pdf",
                "Workforce_Report_20250913_141055.xlsx", 
                "Department_Analysis_20250913_135422.xlsx",
                "Skills_Matrix_20250913_134018.xlsx"
            ]
            
            for export in recent_exports:
                st.text(f"📎 {export}")
